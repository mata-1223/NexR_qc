import os, re, json, copy, time, datetime, traceback
import unicodedata
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from utils.Logging import *
from utils.Timer import *

class QualityCheck:
    
    def __init__(self):
        # 초기 디렉토리 세팅
        self.PATH={}
        self.PATH['ROOT']=os.getcwd()
        self.PATH['DATA']=os.path.join(self.PATH['ROOT'], 'data')
        self.PATH['DOCS']=os.path.join(self.PATH['ROOT'], 'documents')
        self.PATH['LOG']=os.path.join(self.PATH['ROOT'], 'log')
        self.PATH['OUTPUT']=os.path.join(self.PATH['ROOT'], 'output')
        
        for folder in self.PATH.keys():
            # 구성 폴더가 없을 경우, 생성
            if not os.path.exists(self.PATH[folder]):
                Path(self.PATH[folder]).mkdir(parents=True, exist_ok=True)
                
        # 로그 구성
        self.logger_save=False # 로그 파일 생성 여부 (True: 로그 파일 생성 / False: 로그 파일 미생성)
        self.logger=Logger(proc_name='Quality Check', log_folder_path = self.PATH['LOG'], save=self.logger_save)
        
        self.timer=Timer(logger=self.logger)
        self.timer.start()
        
        # Config 파일 불러오기
        with open(os.path.join(self.PATH['ROOT'], 'config.json'), 'r') as f:
            self.config=json.load(f)
        
        self.readFunc={}
        self.readFunc['.csv']=pd.read_csv
        self.readFunc['.xlsx']=pd.read_excel
        
    def data_check(self):
        self.logger.info("="*50)
        self.logger.info("[Step 1] 데이터 파일 존재 여부 확인 시작")
        
        # 데이터 존재 여부 확인
        self.files={os.path.splitext(file)[0].upper(): os.path.join(self.PATH['DATA'], file) for file in os.listdir(self.PATH['DATA']) if not file.startswith('.')}
        
        # 데이터 파일이 존재하지 않을 경우 에러 로그 기록
        if len(self.files.keys())==0:
            self.logger.error(f"QC를 수행할 데이터 파일이 존재하지 않습니다.")
            return
            
        self.logger.info(f"총 {len(self.files):,} 개의 데이터 파일이 존재합니다.")
        self.logger.info(f"{self.files}")
        
        self.DataDict={}
        for name, path in self.files.items():
            self.DataDict[name]={}
            self.DataDict[name]['PATH'] = path
            self.DataDict[name]['EXT'] = os.path.splitext(os.path.basename(path))[-1]
            self.DataDict[name]['DATA'] = self.readFunc[self.DataDict[name]['EXT']](self.DataDict[name]['PATH'], na_values=self.config['naList'], date_format='mixed') # 데이터
        
        self.logger.info("[Step 1] 데이터 파일 존재 여부 확인 완료")
        
    def document_check(self):
        self.logger.info("="*50)
        self.logger.info("[Step 2] 정의서 파일 존재 여부 확인 시작")
        
        # 정의서 파일 존재 여부 확인
        DocList = ['테이블정의서', '컬럼정의서', '코드정의서']
        self.DocumentDict = {}
        for Doc in DocList:
            self.DocumentDict[Doc] = {}
            
            # 존재여부/파일경로 확인
            if len([os.path.join(self.PATH['DOCS'], file) for file in os.listdir(self.PATH['DOCS']) if unicodedata.normalize('NFC', Doc) in unicodedata.normalize('NFC', file)]) > 0:
                self.DocumentDict[Doc]['EXIST'] = True
                self.DocumentDict[Doc]['PATH'] = [os.path.join(self.PATH['DOCS'], file) for file in os.listdir(self.PATH['DOCS']) if unicodedata.normalize('NFC', Doc) in unicodedata.normalize('NFC', file)][0]
                self.DocumentDict[Doc]['EXT'] = os.path.splitext(self.DocumentDict[Doc]['PATH'])[-1]
                self.DocumentDict[Doc]['DATA'] = self.readFunc[self.DocumentDict[Doc]['EXT']](self.DocumentDict[Doc]['PATH'], header = 1) # 데이터
                self.logger.info(f"[{Doc}] 참고할 문서 파일 경로: {self.DocumentDict[Doc]['PATH']}")
            else:
                self.DocumentDict[Doc]['EXIST'] = False
                self.DocumentDict[Doc]['PATH'] = None
                self.DocumentDict[Doc]['EXT'] = None
                self.DocumentDict[Doc]['DATA'] = None
                self.logger.info(f"[{Doc}] 참고할 문서 파일이 없습니다.")
                
        self.logger.info("[Step 2] 정의서 파일 존재 여부를 확인 완료")
        
    def dtype_check(self):
        # 데이터 타입 확인 기능
        
        pass
        
    def na_check(self):
        self.logger.info("="*50)
        self.logger.info("[Step 3] 사전에 등록된 결측값 확인 시작")
        
        # 결측값 Custom 기능
        self.naList = self.config['naList']
        self.logger.info(f"현재 결측값으로 등록된 값은 다음과 같습니다.")
        self.logger.info(f"{self.naList}")
        self.logger.info(f"결측값 추가 등록을 원하시면 config.json 파일 내 naList 값에 추가 시 반영됩니다.")
        
        self.logger.info("[Step 3] 사전에 등록된 결측값 확인 완료")
        
    def run(self):
        
        # QC 수행
        self.logger.info("=" * 50)
        self.logger.info(f"데이터 QC를 수행합니다.")
        
        self.ResultDict = {data_name: {} for data_name in self.DataDict.keys()}
        self.InfoDict = {data_name: {'Table': {}, 'Column': {}} for data_name in self.DataDict.keys()}
        
        # Step 1: 데이터 정보 확인
        self.logger.info('-' * 30)
        self.logger.info(f"[Step 4-1] 데이터 파일 정보 확인 시작")
        
        for i, data_name in enumerate(self.DataDict.keys()):
            self.logger.info(f"{i + 1} 번째 데이터명: {data_name}")
            
        self.logger.info(f"[Step 4-1] 데이터 파일 정보 확인 완료")
        
        # Step 2: 정의서 문서 파일 불러오기
        self.logger.info('-' * 30)
        self.logger.info(f"[Step 4-2] 정의서 문서 정보 확인 시작")
        
        for i, data_name in enumerate(self.DataDict.keys()):
            data = self.DataDict[data_name]['DATA']
            
            # 테이블 정의서 정보 획득
            self.logger.info(f"테이블 정의서 정보 확인 시작")
            if self.DocumentDict['테이블정의서']['EXIST'] == True:
                table_document = self.DocumentDict['테이블정의서']['DATA']
                if data_name in table_document['테이블 영문명'].values:
                    self.logger.info(f'테이블 정의서에 {data_name} 테이블 정보가 존재합니다.')
                    self.InfoDict[data_name]['Table']['스키마명'] = table_document.loc[table_document['테이블 영문명'] == data_name, '스키마명'].values[0]
                    self.InfoDict[data_name]['Table']['테이블 영문명'] = table_document.loc[table_document['테이블 영문명'] == data_name, '테이블 영문명'].values[0]
                    self.InfoDict[data_name]['Table']['테이블 한글명'] = table_document.loc[table_document['테이블 영문명'] == data_name, '테이블 한글명'].values[0]
                else:
                    self.logger.info(f'테이블 정의서에 {data_name} 테이블 정보가 존재하지 않습니다.')
                    self.InfoDict[data_name]['Table']['스키마명'] = None
                    self.InfoDict[data_name]['Table']['테이블 영문명'] = data_name
                    self.InfoDict[data_name]['Table']['테이블 한글명'] = None
                self.InfoDict[data_name]['Table']['테이블 용량'] = None
                self.InfoDict[data_name]['Table']['테이블 기간'] = None
                self.InfoDict[data_name]['Table']['테이블 크기'] = data.shape
                
                self.ResultDict[data_name]['Top'] = pd.DataFrame(self.InfoDict[data_name]['Table'].values(), index=[['스키마명', '테이블 영문명', '테이블 한글명', '테이블 상세', '테이블 상세', '테이블 상세'], ['스키마명', '테이블 영문명', '테이블 한글명', '테이블 용량', '테이블 기간', '테이블 크기']])
                
            else:
                self.logger.info(f'테이블 정의서 문서가 존재하지 않습니다.')
                self.InfoDict[data_name]['Table']['스키마명'] = None
                self.InfoDict[data_name]['Table']['테이블 영문명'] = data_name
                self.InfoDict[data_name]['Table']['테이블 한글명'] = None
                self.InfoDict[data_name]['Table']['테이블 용량'] = None
                self.InfoDict[data_name]['Table']['테이블 기간'] = None
                self.InfoDict[data_name]['Table']['테이블 크기'] = data.shape
                
                self.ResultDict[data_name]['Top'] = pd.DataFrame(self.InfoDict[data_name]['Table'].values(), index=[['스키마명', '테이블 영문명', '테이블 한글명', '테이블 상세', '테이블 상세', '테이블 상세'], ['스키마명', '테이블 영문명', '테이블 한글명', '테이블 용량', '테이블 기간', '테이블 크기']])
                
            self.logger.info(f"테이블 정의서 정보 확인 완료")
            
            # 컬럼 정의서 & 코드 정의서 정보 획득
            self.logger.info(f"코드 정의서 정보 확인 시작")
            if self.DocumentDict['코드정의서']['EXIST'] == True:
                code_document = self.DocumentDict['코드정의서']['DATA']
                code_document.columns=map(lambda x: x.replace('\n', ' '), list(code_document.columns))
                code_values = [val for val in code_document.loc[:, '코드 대분류'].drop_duplicates().values if val is not np.nan]
            else:
                code_document = None
                code_values = None
            
            self.logger.info(f"코드 정의서 정보 확인 완료")
            
            self.logger.info(f"컬럼 정의서 정보 확인 시작")
            for col in data.columns:
                self.InfoDict[data_name]['Column'][col] = {}
                
                # 컬럼 정보 초기 세팅
                self.InfoDict[data_name]['Column'][col]['컬럼 영문명'] = col
                self.InfoDict[data_name]['Column'][col]['컬럼 한글명'] = None
                self.InfoDict[data_name]['Column'][col]['데이터 타입'] = data[col].dtypes.name
                self.InfoDict[data_name]['Column'][col]['코드대분류'] = None
                self.InfoDict[data_name]['Column'][col]['코드값'] = None
            
            if self.DocumentDict['컬럼정의서']['EXIST'] == True:
                column_document = self.DocumentDict['컬럼정의서']['DATA']
                column_document.columns = map(lambda x: x.replace('\n', ' '), list(column_document.columns))
                column_document = column_document.loc[column_document['테이블 영문명'] == data_name, :]
                
                if data_name in column_document['테이블 영문명'].values:
                    self.logger.info(f'컬럼 정의서에 {data_name} 테이블의 컬럼 정보가 존재합니다.')
                    for col in data.columns:
                        if col.upper() in column_document.loc[:, '컬럼 영문명'].values:
                            self.logger.info(f'컬럼 정의서에 {col} 컬럼 정보가 존재합니다.')
                            self.InfoDict[data_name]['Column'][col]['컬럼 영문명'] = col
                            self.InfoDict[data_name]['Column'][col]['컬럼 한글명'] = column_document.loc[column_document['컬럼 영문명'] == col.upper(), '컬럼 한글명'].values[0]
                            self.InfoDict[data_name]['Column'][col]['데이터 타입'] = column_document.loc[column_document['컬럼 영문명'] == col.upper(), '데이터 타입'].values[0]
                            self.InfoDict[data_name]['Column'][col]['코드대분류'] = column_document.loc[column_document['컬럼 영문명'] == col.upper(), '코드대분류'].values[0]
                            self.InfoDict[data_name]['Column'][col]['코드값'] = [val for val in code_document.loc[(code_document.loc[:, '코드 대분류'] == self.InfoDict[data_name]['Column'][col]['코드대분류']), '코드값'].values if val is not np.nan] if self.InfoDict[data_name]['Column'][col]['코드대분류'] in code_values else None
                        else:
                            self.logger.info(f'컬럼 정의서에 {col} 컬럼 정보가 존재하지 않습니다.')
                else: self.logger.info(f'컬럼 정의서에 {data_name} 테이블의 컬럼 정보가 존재하지 않습니다.')
            else: self.logger.info(f'컬럼 정의서 문서가 존재하지 않습니다.')
                        
            self.logger.info(f"[Step 4-2] 정의서 문서 정보 확인 완료")
            
            # Step 3: 결과 항목 값 세팅        
            self.RelCategory={'공통': ['No', '컬럼 영문명', '컬럼 한글명', '데이터 타입', 'null 개수', '%null', '적재건수', '%적재건수'],
                        '연속형': ['최솟값', '최댓값', '평균', '표준편차', '중위수'],
                        '범주형': ['범주수', '범주', '%범주', '정의된 범주 외', '정의된 범주 외 수', '최빈값', '최빈값 수', '%최빈값'],
                        '비고': ['비고']}
            
            self.InfoDict[data_name]['Result'] = {f'{idx+1:03d}': {'공통': {'No': f'{idx+1:03d}', '컬럼 영문명': col}} for idx, col in enumerate(data.columns)}
            
            # Step 4-1: 공통 영역 QC 수행
            for idx in self.InfoDict[data_name]['Result'].keys():
                col = self.InfoDict[data_name]['Result'][idx]['공통']['컬럼 영문명']

                self.InfoDict[data_name]['Result'][idx]['공통']['컬럼 한글명'] = self.InfoDict[data_name]['Column'][col]['컬럼 한글명']  # 컬럼 한글명
                self.InfoDict[data_name]['Result'][idx]['공통']['데이터 타입'] = self.InfoDict[data_name]['Column'][col]['데이터 타입']  # 데이터 타입
                self.InfoDict[data_name]['Result'][idx]['공통']['null 개수'] = '{:,}'.format(data[col].isnull().sum())  # null 개수
                self.InfoDict[data_name]['Result'][idx]['공통']['%null'] = '{:.2%}'.format(data[col].isnull().sum()/data.shape[0])  # %null
                self.InfoDict[data_name]['Result'][idx]['공통']['적재건수'] = '{:,}'.format(data[col].notnull().sum())  # 적재건수
                self.InfoDict[data_name]['Result'][idx]['공통']['%적재건수'] = '{:.2%}'.format(data[col].notnull().sum()/data.shape[0])  # %적재건수
            
            # Step 4-2: 연속형 영역 QC 수행
            for idx in self.InfoDict[data_name]['Result'].keys():
                self.InfoDict[data_name]['Result'][idx]['연속형'] = {}
                col = self.InfoDict[data_name]['Result'][idx]['공통']['컬럼 영문명']
                
                # 연속형 영역 초기값
                self.InfoDict[data_name]['Result'][idx]['연속형']['최솟값'] = None  # 최솟값
                self.InfoDict[data_name]['Result'][idx]['연속형']['최댓값'] = None  # 최댓값
                self.InfoDict[data_name]['Result'][idx]['연속형']['평균'] = None  # 평균
                self.InfoDict[data_name]['Result'][idx]['연속형']['표준편차'] = None  # 표준편차
                self.InfoDict[data_name]['Result'][idx]['연속형']['중위수'] = None  # 표준편차
                
                if any(keyword in self.InfoDict[data_name]['Result'][idx]['공통']['데이터 타입'] for keyword in ['float', 'int', 'numeric']):
                    self.InfoDict[data_name]['Result'][idx]['연속형']['최솟값'] = str(data[col].min())  # 최솟값
                    self.InfoDict[data_name]['Result'][idx]['연속형']['최댓값'] = str(data[col].max())  # 최댓값
                    self.InfoDict[data_name]['Result'][idx]['연속형']['평균'] = str(data[col].mean())  # 평균
                    self.InfoDict[data_name]['Result'][idx]['연속형']['표준편차'] = str(data[col].std())  # 표준편차
                    self.InfoDict[data_name]['Result'][idx]['연속형']['중위수'] = str(np.nanmedian(data[col]))  # 표준편차
                    
                elif any(keyword in self.InfoDict[data_name]['Result'][idx]['공통']['데이터 타입'] for keyword in ['datetime']):
                    self.InfoDict[data_name]['Result'][idx]['연속형']['최솟값'] = str(data[col].min())  # 최솟값
                    self.InfoDict[data_name]['Result'][idx]['연속형']['최댓값'] = str(data[col].max())  # 최댓값
            
            # Step 4-3: 범주형 영역 QC 수행
            for idx in self.InfoDict[data_name]['Result'].keys():
                self.InfoDict[data_name]['Result'][idx]['범주형']={}
                col=self.InfoDict[data_name]['Result'][idx]['공통']['컬럼 영문명']
                
                # 범주형 영역 초기값
                self.InfoDict[data_name]['Result'][idx]['범주형']['범주수'] = None  # 범주수
                self.InfoDict[data_name]['Result'][idx]['범주형']['범주'] = None  # 범주
                self.InfoDict[data_name]['Result'][idx]['범주형']['%범주'] = None  # %범주
                self.InfoDict[data_name]['Result'][idx]['범주형']['정의된 범주 외'] = None  # 정의된 범주 외
                self.InfoDict[data_name]['Result'][idx]['범주형']['정의된 범주 외 수'] = None  # 정의된 범주 외 수
                self.InfoDict[data_name]['Result'][idx]['범주형']['최빈값'] = None  # 최빈값
                self.InfoDict[data_name]['Result'][idx]['범주형']['최빈값 수'] = None  # 최빈값 수
                self.InfoDict[data_name]['Result'][idx]['범주형']['%최빈값'] = None  # %최빈값
                
                
                if any(keyword in self.InfoDict[data_name]['Result'][idx]['공통']['데이터 타입'] for keyword in ['object', 'char', 'varchar', 'datetime']):
                    self.InfoDict[data_name]['Result'][idx]['범주형']['범주수']='{:,}'.format(data[col].nunique(dropna=True)) # 범주수
                    
                    if data[col].nunique(dropna=True) <= 5:
                        self.InfoDict[data_name]['Result'][idx]['범주형']['범주']=data[col].unique().tolist() # 범주
                        self.InfoDict[data_name]['Result'][idx]['범주형']['%범주']={value_: '{:.3%}'.format((data[col].loc[data[col]==value_].shape[0])/(data.shape[0])) for value_ in data[col].unique().tolist()} # %범주
                    else:
                        self.InfoDict[data_name]['Result'][idx]['범주형']['범주']=data[col].unique()[:2].tolist() + ['...'] + data[col].unique()[-2:].tolist() # 범주
                        self.InfoDict[data_name]['Result'][idx]['범주형']['%범주']={value_: '{:.3%}'.format((data[col].loc[data[col]==value_].shape[0])/(data.shape[0])) for value_ in data[col].unique()[:5].tolist()} # %범주
                        self.InfoDict[data_name]['Result'][idx]['범주형']['%범주']['그 외']='{:.3%}'.format((data[col].loc[~(data[col].isin(data[col].unique()[:5].tolist()))].shape[0])/(data.shape[0]))
                        
                    if self.InfoDict[data_name]['Column'][col]['코드값'] is not None:
                        _=[val for val in self.InfoDict[data_name]['Result'][idx]['범주형']['범주'] if val not in self.InfoDict[data_name]['Column'][col]['코드값']]
                        if len(_) > 5:
                            self.InfoDict[data_name]['Result'][idx]['범주형']['정의된 범주 외']=_[:2] + ['...'] + _[-2:]
                        elif len(_) < 1:
                            self.InfoDict[data_name]['Result'][idx]['범주형']['정의된 범주 외']=None
                        else:
                            self.InfoDict[data_name]['Result'][idx]['범주형']['정의된 범주 외']=_
                        self.InfoDict[data_name]['Result'][idx]['범주형']['정의된 범주 외 수']=len(_)
                        
                    if len(data[col].mode(dropna=True).values.tolist()) <= 3:
                        self.InfoDict[data_name]['Result'][idx]['범주형']['최빈값']=data[col].mode(dropna=True).values.tolist() # 최빈값
                        self.InfoDict[data_name]['Result'][idx]['범주형']['최빈값 수']={mode_: '{:,}'.format(data[col].loc[data[col]==mode_].shape[0]) for mode_ in data[col].mode(dropna=True).values.tolist()} # 최빈값 수
                        self.InfoDict[data_name]['Result'][idx]['범주형']['%최빈값']={mode_: '{:.2%}'.format((data[col].loc[data[col]==mode_].shape[0])/(data.shape[0])) for mode_ in data[col].mode(dropna=True).values.tolist()} # %최빈값
                    else:
                        self.InfoDict[data_name]['Result'][idx]['범주형']['최빈값']=data[col].mode(dropna=True).values.tolist()[:2] + ['...'] # 최빈값
                        self.InfoDict[data_name]['Result'][idx]['범주형']['최빈값 수']={mode_col: '{:,}'.format(data[col].loc[data[col]==mode_col].shape[0]) for mode_col in data[col].mode(dropna=True).values.tolist()[:2]} # 최빈값 수
                        self.InfoDict[data_name]['Result'][idx]['범주형']['%최빈값']={mode_col: '{:.2%}'.format((data[col].loc[data[col]==mode_col].shape[0])/(data.shape[0])) for mode_col in data[col].mode(dropna=True).values.tolist()[:2]} # %최빈값

            # Step 4-4: 비고 영역 QC 수행
            for idx in self.InfoDict[data_name]['Result'].keys():
                self.InfoDict[data_name]['Result'][idx]['비고'] = {}
                col = self.InfoDict[data_name]['Result'][idx]['공통']['컬럼 영문명']

                self.InfoDict[data_name]['Result'][idx]['비고']['비고'] = None  # 비고
                
    def convert_to_richtext(self, src):
        if type(src) is list:
            try:
                tgt=',\n'.join(src)
            except TypeError:
                tgt=',\n'.join(map(str, src))
        elif type(src) is dict:
            try:
                tgt="\n".join("{}: {},".format(k, v) for k, v in src.items())[:-1]
            except TypeError:
                tgt=',\n'.join(map(str, src))
            
        return tgt
        
        
    def save(self):
        # 결과 저장
        
        # Step 5: QC 결과서 산출물 생성
        # Step 5-1: 기본 Excel 파일 생성
        SubCol1, SubCol2=[], []
        for key1 in self.RelCategory.keys():
            SubCol1+=[key1] * len(self.RelCategory[key1])
            SubCol2+=self.RelCategory[key1]
        
        ColList=[SubCol1, SubCol2]
        OutputPath=os.path.join(self.PATH['OUTPUT'], 'QC결과서.xlsx')
        
        for data_name in self.DataDict.keys():
            ResultList=[]
            for idx in self.InfoDict[data_name]['Result'].keys():
                ResultList_=[]
                for key1 in self.InfoDict[data_name]['Result'][idx]:
                    ResultList_+=[self.convert_to_richtext(self.InfoDict[data_name]['Result'][idx][key1][key2]) if ((type(self.InfoDict[data_name]['Result'][idx][key1][key2]) is list) or (type(self.InfoDict[data_name]['Result'][idx][key1][key2]) is dict)) else self.InfoDict[data_name]['Result'][idx][key1][key2] for key2 in self.InfoDict[data_name]['Result'][idx][key1]]
                
                ResultList.append(ResultList_)
                
            self.ResultDict[data_name]['Bottom'] = pd.DataFrame(ResultList, columns=ColList)
            
        with pd.ExcelWriter(OutputPath, mode='w', engine='openpyxl') as writer:
            for data_name in self.ResultDict.keys():
                self.ResultDict[data_name]['Top'].to_excel(writer, index=True, header=False, sheet_name=data_name, startcol=1, startrow=1)
                self.ResultDict[data_name]['Bottom'].to_excel(writer, index=True, header=True, sheet_name=data_name, startcol=0, startrow=9)
                
        # Step 5-2: 저장한 Excel 파일 서식 편집 
        wb=load_workbook(OutputPath)
        
        for data_name in self.ResultDict.keys():
            ws=wb[data_name]
            
            ws.delete_rows(12)
            ws.delete_cols(1)
            
            for mcr in ws.merged_cells:
                if 1 < mcr.min_col:
                    mcr.shift(col_shift=-1)
                elif 1 <= mcr.max_col:
                    mcr.shrink(right=1)
                    
            thin = Side(border_style="thin", color="000000")
            
            ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2)
            ws.cell(row=1, column=1).value='테이블 정보'
            ws.cell(row=1, column=1).font=Font(bold=True, color="ffffff")
            ws.cell(row=1, column=1).fill=PatternFill("solid", fgColor="000000")
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=1, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            ws.merge_cells(start_row=9, end_row=9, start_column=1, end_column=2)
            ws.cell(row=9, column=1).value='컬럼 정보'
            ws.cell(row=9, column=1).font=Font(bold=True, color="ffffff")
            ws.cell(row=9, column=1).fill=PatternFill("solid", fgColor="000000")
            ws.cell(row=9, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=9, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            for cell_ in ws['A2':'B7']:
                for cell in cell_:
                    cell.fill = PatternFill("solid", fgColor="bfbfbf")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for cell_ in ws['B5':'B7']:
                for cell in cell_:
                    cell.fill = PatternFill("solid", fgColor="d9d9d9")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for cell_ in ws['C2':'C7']:
                for cell in cell_:
                    cell.alignment = Alignment(vertical='center')
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            for i_, row in enumerate(ws.rows):
                for cell_ in row:
                    if i_==9:
                        if cell_.value in ['공통']:
                            cell=ws[cell_.coordinate]
                            cell.fill = PatternFill("solid", fgColor="bfbfbf")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif cell_.value in ['연속형']:
                            cell=ws[cell_.coordinate]
                            cell.fill = PatternFill("solid", fgColor="f4b084")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif cell_.value in ['범주형']:
                            cell=ws[cell_.coordinate]
                            cell.fill = PatternFill("solid", fgColor="9bc2e6")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif cell_.value in ['비고']:
                            ws.merge_cells(f'V1:V2')
                            cell=ws[cell_.coordinate]
                            cell.fill = PatternFill("solid", fgColor="bfbfbf")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif i_==10:
                        if cell_.value in self.RelCategory['공통'] + self.RelCategory['비고']:
                            cell=ws[cell_.coordinate]
                            cell.fill = PatternFill("solid", fgColor="d9d9d9")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif cell_.value in self.RelCategory['연속형']:
                            cell=ws[cell_.coordinate]
                            cell.fill = PatternFill("solid", fgColor="f8cbad")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif cell_.value in self.RelCategory['범주형']:
                            cell=ws[cell_.coordinate]
                            cell.fill = PatternFill("solid", fgColor="bdd7ee")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif i_>=11:
                        cell=ws[cell_.coordinate]
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            ColumnDimension(ws, bestFit=True)
        
        wb.save(OutputPath)
        
        self.timer.stop()
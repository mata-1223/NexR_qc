import copy
import datetime
import json
import os
import re
import time
import traceback
import unicodedata
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from NexR_qc.Logging import *
from NexR_qc.Timer import *


class QualityCheck:

    def __init__(self, DataDict):
        # DataDict (dict): {'데이터명1': dataframe1, ...}
        # 초기 디렉토리 세팅
        self.PATH = {}
        self.PATH["ROOT"] = os.getcwd()
        self.PATH["DOCS"] = os.path.join(self.PATH["ROOT"], "documents")
        self.PATH["LOG"] = os.path.join(self.PATH["ROOT"], "log")
        self.PATH["OUTPUT"] = os.path.join(self.PATH["ROOT"], "output")

        self.colorSetting = {"grey": "\x1b[38;20m", "blue": "\033[34m", "green": "\033[32m", "yellow": "\x1b[33;20m", "red": "\x1b[31;20m", "bold_red": "\x1b[31;1m", "reset": "\x1b[0m"}

        self.DataDict = {unicodedata.normalize("NFC", k): v for k, v in DataDict.items()}

        for folder in self.PATH.keys():
            # 구성 폴더가 없을 경우, 생성
            if not os.path.exists(self.PATH[folder]):
                Path(self.PATH[folder]).mkdir(parents=True, exist_ok=True)

        # 로그 구성
        self.logger_save = True  # 로그 파일 생성 여부 (True: 로그 파일 생성 / False: 로그 파일 미생성)
        self.logger = Logger(
            proc_name="QualityCheck",
            log_folder_path=self.PATH["LOG"],
            save=self.logger_save,
        )

        self.timer = Timer(logger=self.logger)
        self.timer.start()

        # Config 파일 불러오기
        if "config.json" in os.listdir(self.PATH["ROOT"]):
            with open(os.path.join(self.PATH["ROOT"], "config.json"), "r") as f:
                self.config = json.load(f)
        else:
            self.config = {"naList": ["?", "na", "null", "Null", "NULL", " ", "[NULL]"]}
            with open(os.path.join(self.PATH["ROOT"], "config.json"), "w") as f:
                json.dump(self.config, f)
            self.logger.info(f'config 파일을 생성하였습니다. (생성 경로: {os.path.join(self.PATH["ROOT"], "config.json")})')

        self.readFunc = {}
        self.readFunc[".csv"] = pd.read_csv
        self.readFunc[".xlsx"] = pd.read_excel

    def data_check(self):
        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}[Step 1] 데이터 파일 존재 여부 확인 시작{self.colorSetting['reset']}")

        # 데이터 파일이 존재하지 않을 경우 에러 로그 기록
        if len(self.DataDict.keys()) == 0:
            self.logger.error(f"QC를 수행할 데이터 파일이 존재하지 않습니다.")
            return

        self.logger.info(f"총 {len(self.DataDict):,} 개의 데이터 파일이 존재합니다.")

        # 날짜 혹은 시간 컬럼 관련 추가 정보 입력 필요 여부값 확인
        self.config["DateTimeInfoQuestion_YN"] = True if input(f"""{self.colorSetting['yellow']}각 테이블 내 날짜 혹은 시간 관련 컬럼에 대한 추가 정보 입력이 필요한 경우 Y, 추가 정보 입력이 필요 없는 경우는 N을 입력해주세요 (Y/N):{self.colorSetting['reset']} """) in ["Y", "y"] else False

        for name, data in self.DataDict.items():
            self.DataDict[name] = {}
            self.DataDict[name]["DATA"] = data.replace(self.config["naList"], np.nan)
            self.DataDict[name]["TIMECOL"] = None
            if not self.config["DateTimeInfoQuestion_YN"]:
                self.DataDict[name]["TIMECOL"] = []
            else:
                while True:
                    # 날짜 혹은 시간 컬럼 관련 추가 정보 입력
                    self.logger.info(f"""컬럼 중 날짜 혹은 시간 관련 컬럼 존재 여부를 알려주세요. (컬럼 정의서에 명시가 되어있는 경우나 날짜 혹은 시간 관련 컬럼이 없는 경우는 Enter로 넘어가셔도 됩니다.)\n현재 {name} 데이터의 컬럼은 다음과 같습니다.\n{self.DataDict[name]["DATA"].columns.tolist()}\n\n""")
                    col_ = list(input(f"""{self.colorSetting["yellow"]}[{name}] 날짜 혹은 시간 관련 컬럼:{self.colorSetting["reset"]} """).split(","))
                    self.DataDict[name]["TIMECOL"] = [col.strip() for col in col_] if col_ != [""] else []
                    self.logger.info(f"""{self.colorSetting["yellow"]}[{name}] 날짜 혹은 시간 관련 컬럼:{self.colorSetting["reset"]}: {self.DataDict[name]["TIMECOL"]}""")
                    # 입력값 유효성 검증
                    if all(time_col in self.DataDict[name]["DATA"].columns.tolist() for time_col in self.DataDict[name]["TIMECOL"]):
                        break
                    self.logger.error("⛔️ 컬럼명 입력값이 잘못 입력되었습니다. 입력하신 컬럼명을 다시 한번 확인해주세요.")
        self.logger.info(f"{self.colorSetting['green']}[Step 1] 데이터 파일 존재 여부 확인 완료{self.colorSetting['reset']}")

    def document_check(self):
        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}[Step 2] 정의서 파일 존재 여부 확인 시작{self.colorSetting['reset']}")

        # 정의서 파일 존재 여부 확인
        DocList = ["테이블정의서", "컬럼정의서", "코드정의서"]
        self.DocumentDict = {}
        for Doc in DocList:
            self.DocumentDict[Doc] = {}

            # 존재여부/파일경로 확인
            if len([os.path.join(self.PATH["DOCS"], file) for file in os.listdir(self.PATH["DOCS"]) if unicodedata.normalize("NFC", Doc) in unicodedata.normalize("NFC", file)]) > 0:
                self.DocumentDict[Doc]["EXIST"] = True
                self.DocumentDict[Doc]["PATH"] = [os.path.join(self.PATH["DOCS"], file) for file in os.listdir(self.PATH["DOCS"]) if unicodedata.normalize("NFC", Doc) in unicodedata.normalize("NFC", file)][0]
                self.DocumentDict[Doc]["EXT"] = os.path.splitext(self.DocumentDict[Doc]["PATH"])[-1]
                self.DocumentDict[Doc]["DATA"] = self.readFunc[self.DocumentDict[Doc]["EXT"]](self.DocumentDict[Doc]["PATH"], header=1)  # 데이터
                self.logger.info(f"[{Doc}] 참고할 문서 파일 경로: {self.DocumentDict[Doc]['PATH']}")
            else:
                self.DocumentDict[Doc]["EXIST"] = False
                self.DocumentDict[Doc]["PATH"] = None
                self.DocumentDict[Doc]["EXT"] = None
                self.DocumentDict[Doc]["DATA"] = None
                self.logger.info(f"[{Doc}] 참고할 문서 파일이 없습니다.")

        self.logger.info(f"{self.colorSetting['green']}[Step 2] 정의서 파일 존재 여부를 확인 완료{self.colorSetting['reset']}")

    def na_check(self):
        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}[Step 3] 사전에 등록된 결측값 확인 시작{self.colorSetting['reset']}")

        # 결측값 Custom 기능
        self.naList = self.config["naList"]
        self.logger.info(f"현재 결측값으로 등록된 값은 다음과 같습니다.")
        self.logger.info(f"{self.naList}")
        self.logger.info(f"결측값 추가 등록을 원하시면 config.json 파일 내 naList 값에 추가 시 반영됩니다.")

        self.logger.info(f"{self.colorSetting['green']}[Step 3] 사전에 등록된 결측값 확인 완료{self.colorSetting['reset']}")

    def run(self):

        # QC 수행
        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}[Step 4] QC 사전 정보 확인 시작{self.colorSetting['reset']}")

        self.ResultDict = {data_name: {} for data_name in self.DataDict.keys()}
        self.InfoDict = {data_name: {"Table": {}, "Column": {}} for data_name in self.DataDict.keys()}

        # Step 1: 데이터 정보 확인
        self.logger.info("-" * 30)
        self.logger.info(f"[Step 4-1] 데이터 자체 정보 확인 시작")

        for i, data_name in enumerate(self.DataDict.keys()):
            self.logger.info(f"{i + 1} 번째 데이터명: {data_name}")

        self.logger.info(f"[Step 4-1] 데이터 자체 정보 확인 완료")

        # Step 2: 정의서 문서 파일 불러오기
        self.logger.info("-" * 30)
        self.logger.info(f"[Step 4-2] 정의서 문서 정보 확인 시작")

        for i, data_name in enumerate(self.DataDict.keys()):
            data = self.DataDict[data_name]["DATA"]

            # 테이블 정의서 정보 획득
            self.logger.info(f"테이블 정의서 내 {data_name} 정보 확인 시작")
            if self.DocumentDict["테이블정의서"]["EXIST"] == True:
                table_document = self.DocumentDict["테이블정의서"]["DATA"]
                if data_name in table_document["테이블 영문명"].values:
                    self.logger.info(f"테이블 정의서에 {data_name} 테이블 정보가 존재합니다.")
                    self.InfoDict[data_name]["Table"]["스키마명"] = table_document.loc[table_document["테이블 영문명"] == data_name, "스키마명"].values[0]
                    self.InfoDict[data_name]["Table"]["테이블 영문명"] = table_document.loc[table_document["테이블 영문명"] == data_name, "테이블 영문명"].values[0]
                    self.InfoDict[data_name]["Table"]["테이블 한글명"] = table_document.loc[table_document["테이블 영문명"] == data_name, "테이블 한글명"].values[0]
                else:
                    self.logger.info(f"테이블 정의서에 {data_name} 테이블 정보가 존재하지 않습니다.")
                    self.InfoDict[data_name]["Table"]["스키마명"] = None
                    self.InfoDict[data_name]["Table"]["테이블 영문명"] = data_name
                    self.InfoDict[data_name]["Table"]["테이블 한글명"] = None
                self.InfoDict[data_name]["Table"]["테이블 용량"] = None
                self.InfoDict[data_name]["Table"]["테이블 기간"] = None
                self.InfoDict[data_name]["Table"]["테이블 크기"] = data.shape

                self.ResultDict[data_name]["Top"] = pd.DataFrame(
                    self.InfoDict[data_name]["Table"].values(),
                    index=[
                        ["스키마명", "테이블 영문명", "테이블 한글명", "테이블 상세", "테이블 상세", "테이블 상세"],
                        ["스키마명", "테이블 영문명", "테이블 한글명", "테이블 용량", "테이블 기간", "테이블 크기"],
                    ],
                )

            else:
                self.logger.info(f"테이블 정의서 문서가 존재하지 않습니다.")
                self.InfoDict[data_name]["Table"]["스키마명"] = None
                self.InfoDict[data_name]["Table"]["테이블 영문명"] = data_name
                self.InfoDict[data_name]["Table"]["테이블 한글명"] = None
                self.InfoDict[data_name]["Table"]["테이블 용량"] = None
                self.InfoDict[data_name]["Table"]["테이블 기간"] = None
                self.InfoDict[data_name]["Table"]["테이블 크기"] = data.shape

                self.ResultDict[data_name]["Top"] = pd.DataFrame(
                    self.InfoDict[data_name]["Table"].values(),
                    index=[
                        ["스키마명", "테이블 영문명", "테이블 한글명", "테이블 상세", "테이블 상세", "테이블 상세"],
                        ["스키마명", "테이블 영문명", "테이블 한글명", "테이블 용량", "테이블 기간", "테이블 크기"],
                    ],
                )

            self.logger.info(f"테이블 정의서 내 {data_name} 정보 확인 완료")

            # 컬럼 정의서 & 코드 정의서 정보 획득
            self.logger.info(f"코드 정의서 내 {data_name} 정보 확인 시작")
            if self.DocumentDict["코드정의서"]["EXIST"] == True:
                code_document = self.DocumentDict["코드정의서"]["DATA"]
                code_document.columns = map(lambda x: x.replace("\n", " "), list(code_document.columns))
                code_values = [val for val in code_document.loc[:, "코드 대분류"].drop_duplicates().values if val is not np.nan]
            else:
                code_document = None
                code_values = None

            self.logger.info(f"코드 정의서 내 {data_name} 정보 확인 완료")

            self.logger.info(f"컬럼 정의서 내 {data_name} 정보 확인 시작")
            for col in data.columns:
                self.InfoDict[data_name]["Column"][col] = {}

                # 컬럼 정보 초기 세팅
                self.InfoDict[data_name]["Column"][col]["컬럼 영문명"] = col
                self.InfoDict[data_name]["Column"][col]["컬럼 한글명"] = None
                self.InfoDict[data_name]["Column"][col]["데이터 타입"] = "datetime" if col in self.DataDict[data_name]["TIMECOL"] else data[col].dtypes.name
                self.InfoDict[data_name]["Column"][col]["코드대분류"] = None
                self.InfoDict[data_name]["Column"][col]["코드값"] = None

            if self.DocumentDict["컬럼정의서"]["EXIST"] == True:
                column_document = self.DocumentDict["컬럼정의서"]["DATA"]
                column_document.columns = map(lambda x: x.replace("\n", " "), list(column_document.columns))
                column_document = column_document.loc[column_document["테이블 영문명"] == data_name, :]

                if data_name in column_document["테이블 영문명"].values:
                    self.logger.info(f"컬럼 정의서에 {data_name} 테이블의 컬럼 정보가 존재합니다.")
                    for col in data.columns:
                        if col.upper() in column_document.loc[:, "컬럼 영문명"].values:
                            self.logger.info(f"컬럼 정의서에 {data_name} 테이블의 {col} 컬럼 정보가 존재합니다.")
                            self.InfoDict[data_name]["Column"][col]["컬럼 영문명"] = col
                            self.InfoDict[data_name]["Column"][col]["컬럼 한글명"] = column_document.loc[column_document["컬럼 영문명"] == col.upper(), "컬럼 한글명"].values[0]
                            self.InfoDict[data_name]["Column"][col]["데이터 타입"] = "datetime" if col in self.DataDict[data_name]["TIMECOL"] else column_document.loc[column_document["컬럼 영문명"] == col.upper(), "데이터 타입"].values[0]
                            self.InfoDict[data_name]["Column"][col]["코드대분류"] = column_document.loc[column_document["컬럼 영문명"] == col.upper(), "코드대분류"].values[0]
                            self.InfoDict[data_name]["Column"][col]["코드값"] = [val for val in code_document.loc[(code_document.loc[:, "코드 대분류"] == self.InfoDict[data_name]["Column"][col]["코드대분류"]), "코드값"].values if val is not np.nan] if self.InfoDict[data_name]["Column"][col]["코드대분류"] in code_values else None
                        else:
                            self.logger.info(f"컬럼 정의서에 {data_name} 테이블의 {col} 컬럼 정보가 존재하지 않습니다.")
                else:
                    self.logger.info(f"컬럼 정의서에 {data_name} 테이블의 컬럼 정보가 존재하지 않습니다.")
            else:
                self.logger.info(f"컬럼 정의서 문서가 존재하지 않습니다.")

        # 테이블 리스트 정보 획득
        TableList_ = [[idx + 1, self.InfoDict[data_name]["Table"]["스키마명"], self.InfoDict[data_name]["Table"]["테이블 영문명"], self.InfoDict[data_name]["Table"]["테이블 한글명"], f"{idx+1:04d}_{self.InfoDict[data_name]['Table']['테이블 영문명'][:26]}"] for idx, data_name in enumerate(self.DataDict.keys())]
        self.InfoDict["TableList"] = pd.DataFrame(TableList_, columns=["No.", "스키마명", "테이블 영문명", "테이블 한글명", "워크 시트명"])

        self.logger.info(f"[Step 4-2] 정의서 문서 정보 확인 완료")
        self.logger.info(f"{self.colorSetting['green']}[Step 4] QC 사전 정보 확인 완료{self.colorSetting['reset']}")

        # Step 5. 항목별 QC 실행
        # 결과 항목 값 세팅
        self.RelCategory = {
            "공통": ["No", "컬럼 영문명", "컬럼 한글명", "데이터 타입", "null 개수", "%null", "적재건수", "%적재건수"],
            "연속형": ["최솟값", "최댓값", "평균", "표준편차", "중위수"],
            "범주형": ["범주수", "범주", "%범주", "정의된 범주 외", "정의된 범주 외 수", "최빈값", "최빈값 수", "%최빈값"],
            "비고": ["비고"],
        }

        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}[Step 5] 항목별 데이터 QC 시작{self.colorSetting['reset']}")
        for i, data_name in enumerate(self.DataDict.keys()):
            self.logger.info(f"[{data_name}] QC 시작")
            data = self.DataDict[data_name]["DATA"]

            self.InfoDict[data_name]["Result"] = {f"{idx+1:03d}": {"공통": {"No": f"{idx+1:03d}", "컬럼 영문명": col}} for idx, col in enumerate(data.columns)}

            # QC 항목별 초기값 설정
            for idx in self.InfoDict[data_name]["Result"].keys():

                # 공통 영역 초기값
                self.InfoDict[data_name]["Result"][idx]["공통"]["컬럼 한글명"] = None  # 컬럼 한글명
                self.InfoDict[data_name]["Result"][idx]["공통"]["데이터 타입"] = None  # 데이터 타입
                self.InfoDict[data_name]["Result"][idx]["공통"]["null 개수"] = None  # null 개수
                self.InfoDict[data_name]["Result"][idx]["공통"]["%null"] = None  # %null
                self.InfoDict[data_name]["Result"][idx]["공통"]["적재건수"] = None  # 적재건수
                self.InfoDict[data_name]["Result"][idx]["공통"]["%적재건수"] = None  # %적재건수

                # 연속형 영역 초기값
                self.InfoDict[data_name]["Result"][idx]["연속형"] = {}
                self.InfoDict[data_name]["Result"][idx]["연속형"]["최솟값"] = None  # 최솟값
                self.InfoDict[data_name]["Result"][idx]["연속형"]["최댓값"] = None  # 최댓값
                self.InfoDict[data_name]["Result"][idx]["연속형"]["평균"] = None  # 평균
                self.InfoDict[data_name]["Result"][idx]["연속형"]["표준편차"] = None  # 표준편차
                self.InfoDict[data_name]["Result"][idx]["연속형"]["중위수"] = None  # 표준편차

                # 범주형 영역 초기값
                self.InfoDict[data_name]["Result"][idx]["범주형"] = {}
                self.InfoDict[data_name]["Result"][idx]["범주형"]["범주수"] = None  # 범주수
                self.InfoDict[data_name]["Result"][idx]["범주형"]["범주"] = None  # 범주
                self.InfoDict[data_name]["Result"][idx]["범주형"]["%범주"] = None  # %범주
                self.InfoDict[data_name]["Result"][idx]["범주형"]["정의된 범주 외"] = None  # 정의된 범주 외
                self.InfoDict[data_name]["Result"][idx]["범주형"]["정의된 범주 외 수"] = None  # 정의된 범주 외 수
                self.InfoDict[data_name]["Result"][idx]["범주형"]["최빈값"] = None  # 최빈값
                self.InfoDict[data_name]["Result"][idx]["범주형"]["최빈값 수"] = None  # 최빈값 수
                self.InfoDict[data_name]["Result"][idx]["범주형"]["%최빈값"] = None  # %최빈값

                # 비고 영역 초기값
                self.InfoDict[data_name]["Result"][idx]["비고"] = {}
                self.InfoDict[data_name]["Result"][idx]["비고"]["비고"] = None  # 비고

            # Step 5-1: 공통 영역 QC 수행
            for idx in self.InfoDict[data_name]["Result"].keys():
                col = self.InfoDict[data_name]["Result"][idx]["공통"]["컬럼 영문명"]
                self.InfoDict[data_name]["Result"][idx]["공통"]["컬럼 한글명"] = self.InfoDict[data_name]["Column"][col]["컬럼 한글명"]  # 컬럼 한글명
                self.InfoDict[data_name]["Result"][idx]["공통"]["데이터 타입"] = self.InfoDict[data_name]["Column"][col]["데이터 타입"]  # 데이터 타입
                self.InfoDict[data_name]["Result"][idx]["공통"]["null 개수"] = "{:,}".format(data[col].isnull().sum())  # null 개수
                self.InfoDict[data_name]["Result"][idx]["공통"]["%null"] = "{:.2%}".format(data[col].isnull().sum() / data.shape[0])  # %null
                self.InfoDict[data_name]["Result"][idx]["공통"]["적재건수"] = "{:,}".format(data[col].notnull().sum())  # 적재건수
                self.InfoDict[data_name]["Result"][idx]["공통"]["%적재건수"] = "{:.2%}".format(data[col].notnull().sum() / data.shape[0])  # %적재건수

                # 모든 값이 결측값인 경우, 비고에 알림 문구 작성
                if data[col].isnull().sum() == data.shape[0]:
                    self.InfoDict[data_name]["Result"][idx]["비고"]["비고"] = "결측값 100%"  # 비고

                try:

                    # Step 5-2: 연속형 영역 QC 수행
                    if any(keyword in self.InfoDict[data_name]["Result"][idx]["공통"]["데이터 타입"] for keyword in ["float", "int", "numeric"]):
                        self.InfoDict[data_name]["Result"][idx]["연속형"]["최솟값"] = str(data[col].min())  # 최솟값
                        self.InfoDict[data_name]["Result"][idx]["연속형"]["최댓값"] = str(data[col].max())  # 최댓값
                        self.InfoDict[data_name]["Result"][idx]["연속형"]["평균"] = str(data[col].mean())  # 평균
                        self.InfoDict[data_name]["Result"][idx]["연속형"]["표준편차"] = str(data[col].std())  # 표준편차
                        self.InfoDict[data_name]["Result"][idx]["연속형"]["중위수"] = str(np.nanmedian(data[col]))  # 표준편차

                    elif any(keyword in self.InfoDict[data_name]["Result"][idx]["공통"]["데이터 타입"] for keyword in ["datetime"]):
                        self.InfoDict[data_name]["Result"][idx]["연속형"]["최솟값"] = str(data[col].min())  # 최솟값
                        self.InfoDict[data_name]["Result"][idx]["연속형"]["최댓값"] = str(data[col].max())  # 최댓값

                    # Step 5-3: 범주형 영역 QC 수행
                    colData = data[col].dropna()

                    if any(keyword in self.InfoDict[data_name]["Result"][idx]["공통"]["데이터 타입"] for keyword in ["object", "char", "varchar", "datetime"]):
                        self.InfoDict[data_name]["Result"][idx]["범주형"]["범주수"] = "{:,}".format(colData.nunique(dropna=True))  # 범주수

                        if colData.nunique(dropna=True) <= 5:
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["범주"] = colData.dropna().unique().tolist()  # 범주
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["%범주"] = {value_: "{:.3%}".format((colData.loc[colData == value_].shape[0]) / (colData.shape[0])) for value_ in colData.unique().tolist()}  # %범주
                        else:
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["범주"] = colData.unique()[:2].tolist() + ["..."] + colData.unique()[-2:].tolist()  # 범주
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["%범주"] = {value_: "{:.3%}".format((colData.loc[colData == value_].shape[0]) / (colData.shape[0])) for value_ in colData.unique()[:5].tolist()}  # %범주
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["%범주"]["그 외"] = "{:.3%}".format((colData.loc[~(colData.isin(colData.unique()[:5].tolist()))].shape[0]) / (colData.shape[0]))

                        if self.InfoDict[data_name]["Column"][col]["코드값"] is not None:
                            _ = [val for val in self.InfoDict[data_name]["Result"][idx]["범주형"]["범주"] if val not in self.InfoDict[data_name]["Column"][col]["코드값"]]
                            if len(_) > 5:
                                self.InfoDict[data_name]["Result"][idx]["범주형"]["정의된 범주 외"] = _[:2] + ["..."] + _[-2:]
                            elif len(_) < 1:
                                self.InfoDict[data_name]["Result"][idx]["범주형"]["정의된 범주 외"] = None
                            else:
                                self.InfoDict[data_name]["Result"][idx]["범주형"]["정의된 범주 외"] = _
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["정의된 범주 외 수"] = len(_)

                        if len(colData.mode(dropna=True).values.tolist()) <= 3:
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["최빈값"] = colData.mode(dropna=True).values.tolist()  # 최빈값
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["최빈값 수"] = {mode_: "{:,}".format(colData.loc[colData == mode_].shape[0]) for mode_ in colData.mode(dropna=True).values.tolist()}  # 최빈값 수
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["%최빈값"] = {mode_: "{:.2%}".format((colData.loc[colData == mode_].shape[0]) / (colData.shape[0])) for mode_ in colData.mode(dropna=True).values.tolist()}  # %최빈값
                        else:
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["최빈값"] = colData.mode(dropna=True).values.tolist()[:2] + ["..."]  # 최빈값
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["최빈값 수"] = {mode_col: "{:,}".format(colData.loc[colData == mode_col].shape[0]) for mode_col in colData.mode(dropna=True).values.tolist()[:2]}  # 최빈값 수
                            self.InfoDict[data_name]["Result"][idx]["범주형"]["%최빈값"] = {mode_col: "{:.2%}".format((colData.loc[colData == mode_col].shape[0]) / (colData.shape[0])) for mode_col in colData.mode(dropna=True).values.tolist()[:2]}  # %최빈값

                except TypeError:
                    # 컬럼정의서 데이터 형식과 실데이터 형식 불일치할 경우
                    self.logger.error(f"{data_name} 테이블의 {col} 컬럼 에러")
                    self.logger.error(traceback.format_exc())

                    self.InfoDict[data_name]["Result"][idx]["비고"]["비고"] = "컬럼 정의서 상의 데이터 타입과 실제 데이터 타입 불일치"  # 비고

                pass

            self.logger.info(f"[{data_name}] QC 완료")

        self.logger.info(f"{self.colorSetting['green']}[Step 5] 항목별 데이터 QC 완료{self.colorSetting['reset']}")

    def convert_to_richtext(self, src):
        if type(src) is list:
            try:
                tgt = ",\n".join(src)
            except TypeError:
                tgt = ",\n".join(map(str, src))
        elif type(src) is dict:
            try:
                tgt = "\n".join("{}: {},".format(k, v) for k, v in src.items())[:-1]
            except TypeError:
                tgt = ",\n".join(map(str, src))

        return tgt

    def save(self):
        # 결과 저장

        # Step 6: QC 결과서 산출물 생성
        # Step 6-1: 기본 Excel 파일 생성
        self.logger.info("=" * 50)
        self.logger.info(f"{self.colorSetting['green']}[Step 6] 데이터 QC 결과 저장 작업 시작{self.colorSetting['reset']}")
        SubCol1, SubCol2 = [], []
        for key1 in self.RelCategory.keys():
            SubCol1 += [key1] * len(self.RelCategory[key1])
            SubCol2 += self.RelCategory[key1]

        ColList = [SubCol1, SubCol2]

        OutputCreatedTime = datetime.today()
        OutputCreatedTime = OutputCreatedTime.strftime("%Y%m%d_%H%M%S")
        OutputPath = os.path.join(self.PATH["OUTPUT"], f"QC결과서_{OutputCreatedTime}.xlsx")

        for data_name in self.DataDict.keys():
            ResultList = []
            for idx in self.InfoDict[data_name]["Result"].keys():
                ResultList_ = []
                for key1 in self.InfoDict[data_name]["Result"][idx]:
                    ResultList_ += [(self.convert_to_richtext(self.InfoDict[data_name]["Result"][idx][key1][key2]) if ((type(self.InfoDict[data_name]["Result"][idx][key1][key2]) is list) or (type(self.InfoDict[data_name]["Result"][idx][key1][key2]) is dict)) else self.InfoDict[data_name]["Result"][idx][key1][key2]) for key2 in self.InfoDict[data_name]["Result"][idx][key1]]

                ResultList.append(ResultList_)

            self.ResultDict[data_name]["Bottom"] = pd.DataFrame(ResultList, columns=ColList)

        with pd.ExcelWriter(OutputPath, mode="w", engine="openpyxl") as writer:

            # Step 6-1-a 테이블 리스트 시트
            self.InfoDict["TableList"].to_excel(
                writer,
                index=False,
                header=True,
                sheet_name="테이블 리스트",
                startcol=0,
                startrow=0,
            )

            self.logger.info(f"테이블 리스트 시트 생성 완료")

            # Step 6-1-a 테이블 별 QC 결과서 시트
            for idx, data_name in enumerate(self.ResultDict.keys()):
                self.ResultDict[data_name]["Top"].to_excel(
                    writer,
                    index=True,
                    header=False,
                    sheet_name=f"{idx+1:04d}_{self.InfoDict[data_name]['Table']['테이블 영문명'][:26]}",
                    startcol=1,
                    startrow=1,
                )
                self.ResultDict[data_name]["Bottom"].to_excel(
                    writer,
                    index=True,
                    header=True,
                    sheet_name=f"{idx+1:04d}_{self.InfoDict[data_name]['Table']['테이블 영문명'][:26]}",
                    startcol=0,
                    startrow=9,
                )

                if any([(idx + 1) % 10 == 0, (idx + 1) == len(self.ResultDict.keys())]):
                    self.logger.info(f"{idx + 1} / {len(self.ResultDict.keys())} 번째 엑셀 시트 생성 완료")

        # Step 6-2: 저장한 Excel 파일 서식 편집
        wb = load_workbook(OutputPath)

        thin = Side(border_style="thin", color="000000")

        # Step 6-2-a 테이블 리스트 시트
        ws = wb["테이블 리스트"]

        for cell_ in ws["A1":"D1"]:
            for cell in cell_:
                cell.fill = PatternFill("solid", fgColor="ededed")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for row_i_, row in enumerate(ws.rows):
            for col_i_, cell_ in enumerate(row):
                if all([row_i_ > 0, col_i_ == 0]):
                    cell = ws[cell_.coordinate]
                    cell.fill = PatternFill("solid", fgColor="d8d8d8")
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                elif all([row_i_ > 0, col_i_ != 2]):
                    cell = ws[cell_.coordinate]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                elif all([row_i_ > 0, col_i_ == 2]):
                    cell = ws[cell_.coordinate]
                    hyperlink_value = f"#'{ws.cell(row = row_i_+1, column = col_i_ + 3).value}'!A1"
                    cell.hyperlink = f"{hyperlink_value}"
                    cell.style = "Hyperlink"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        ws.column_dimensions["A"].width = 13.67
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 24.33
        ws.column_dimensions["D"].width = 60.83

        ws.delete_cols(5)
        self.logger.info(f"테이블 리스트 시트 서식 편집 완료")

        # Step 6-2-b 테이블 별 QC 결과서 시트
        for idx, data_name in enumerate(self.ResultDict.keys()):

            try:

                sheet_name = f"{idx+1:04d}_{self.InfoDict[data_name]['Table']['테이블 영문명'][:26]}"
                ws = wb[sheet_name]

                ws.delete_rows(12)
                ws.delete_cols(1)

                for mcr in ws.merged_cells:
                    if 1 < mcr.min_col:
                        mcr.shift(col_shift=-1)
                    elif 1 <= mcr.max_col:
                        mcr.shrink(right=1)

                ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=2)
                ws.cell(row=1, column=1).value = "테이블 정보"
                ws.cell(row=1, column=1).font = Font(bold=True, color="ffffff")
                ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="000000")
                ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=1, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=2)
                ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=2)
                ws.merge_cells(start_row=4, end_row=4, start_column=1, end_column=2)

                ws.merge_cells(start_row=9, end_row=9, start_column=1, end_column=2)
                ws.cell(row=9, column=1).value = "컬럼 정보"
                ws.cell(row=9, column=1).font = Font(bold=True, color="ffffff")
                ws.cell(row=9, column=1).fill = PatternFill("solid", fgColor="000000")
                ws.cell(row=9, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=9, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)

                for cell_ in ws["A2":"B7"]:
                    for cell in cell_:
                        cell.fill = PatternFill("solid", fgColor="bfbfbf")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                for cell_ in ws["B5":"B7"]:
                    for cell in cell_:
                        cell.fill = PatternFill("solid", fgColor="d9d9d9")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                for cell_ in ws["C2":"C7"]:
                    for cell in cell_:
                        cell.alignment = Alignment(vertical="center")
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                for i_, row in enumerate(ws.rows):
                    for cell_ in row:
                        if i_ == 9:
                            if cell_.value in ["공통"]:
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="bfbfbf")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            elif cell_.value in ["연속형"]:
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="f4b084")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            elif cell_.value in ["범주형"]:
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="9bc2e6")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            elif cell_.value in ["비고"]:
                                ws.merge_cells(f"V1:V2")
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="bfbfbf")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif i_ == 10:
                            if cell_.value in self.RelCategory["공통"] + self.RelCategory["비고"]:
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="d9d9d9")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            elif cell_.value in self.RelCategory["연속형"]:
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="f8cbad")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            elif cell_.value in self.RelCategory["범주형"]:
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="bdd7ee")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif i_ >= 11:
                            if cell_.value == "컬럼 정의서 상의 데이터 타입과 실제 데이터 타입 불일치":
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="f79645")
                                cell.font = Font(bold=True, color="ff0000")
                                cell.alignment = Alignment(vertical="center", wrap_text=True)
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            elif cell_.value == "결측값 100%":
                                cell = ws[cell_.coordinate]
                                cell.fill = PatternFill("solid", fgColor="ffff00")
                                cell.font = Font(bold=True, color="ff0000")
                                cell.alignment = Alignment(vertical="center", wrap_text=True)
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                            else:
                                cell = ws[cell_.coordinate]
                                cell.alignment = Alignment(vertical="center", wrap_text=True)
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

                ColumnDimension(ws, bestFit=True)

                if any([(idx + 1) % 10 == 0, (idx + 1) == len(self.ResultDict.keys())]):
                    self.logger.info(f"{idx + 1} / {len(self.ResultDict.keys())} 번째 엑셀 시트 서식 편집 완료")

            except:
                self.logger.error(traceback.format_exc())
                self.logger.error(f"{idx + 1} / {len(self.ResultDict.keys())} 번째 엑셀 시트 서식 편집 실패")
                pass

        wb.save(OutputPath)
        self.logger.info(f"{self.colorSetting['green']}[Step 6] 데이터 QC 결과 저장 작업 완료{self.colorSetting['reset']}")
        self.logger.info(f"{self.colorSetting['green']}모든 QC 프로세스가 완료되었습니다.{self.colorSetting['reset']}")
        self.timer.stop()
        self.logger.info(f"산출물 파일 경로: {self.colorSetting['blue']}{OutputPath}{self.colorSetting['reset']}")
